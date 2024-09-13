import openpyxl
#
# def read_excel(filename, n1, n2):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb['Sheet1']
#     cell = sheet.cell(row = n1, column=n2)
#     print(cell.value)
#
# read_excel('c:\\PythonPractice\\PythonCode\\Santosh\\PythonProgramming\\Python_File_Handling_Operations\\Python_File_Handling1_Excel.xlsx', 1, 5)
#
# def read_excel_with_cellname(filename):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb['Sheet1']
#     cell = sheet['A7']
#     print(cell.value)
#
# read_excel_with_cellname('c:\\PythonPractice\\PythonCode\\Santosh\\PythonProgramming\\Python_File_Handling_Operations\\Python_File_Handling1_Excel.xlsx')
#
#
#
# def read_excel_with_maxrow_maxcol(filename, SheetName):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb[SheetName]
#     rows = sheet.rows
#     for cell in rows:
#         print('Rows', cell[2].value)
#
#     max_colums = sheet.columns
#     for cell in max_colums:
#         print('Column', cell[0].value)
#
# read_excel_with_maxrow_maxcol('c:\\PythonPractice\\PythonCode\\Santosh\\PythonProgramming\\Python_File_Handling_Operations\\Python_File_Handling1_Excel.xlsx', 'Sheet1')
#
#
# #######################
# # update the exel sheet data
#
# def write_excel(filename, n1, n2):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb['Sheet1']
#     sheet['A20'] = n1
#     sheet['A21'] = n2
#     wb.save(filename)
#
#
# write_excel('c:\\PythonPractice\\PythonCode\\Santosh\\PythonProgramming\\Python_File_Handling_Operations\\Python_File_Handling1_Excel.xlsx', 'Santosh', 'Rachotimath')
#
#
# ##Q1 write a python program to enter all even number from 1 to 20 in A colom
#
# def even_numbers(filename, count, i):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb['Sheet2']
#     sheet[f'C{count}'] = i
#
#     wb.save(filename)
#
# count = 0
# for i in range(1, 21):
#     if i % 2 == 0:
#         count += 1
#         even_numbers(
#         'c:\\PythonPractice\\PythonCode\\Santosh\\PythonProgramming\\Python_File_Handling_Operations\\Python_File_Handling1_Excel.xlsx', count, i )
#
# #Q2 write a python program to enter a table of 5 in excel sheet
#
# def tables(filename, num):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb['Sheet3']
#     for i in range(1, 11):
#         sheet[f'F{i}'] = f'{num}*{i}'
#         sheet[f'G{i}'] = num * i
#
#     wb.save(filename)
#
#
# tables('c:\\PythonPractice\\PythonCode\\Santosh\\PythonProgramming\\Python_File_Handling_Operations\\Python_File_Handling1_Excel.xlsx', 5 )
