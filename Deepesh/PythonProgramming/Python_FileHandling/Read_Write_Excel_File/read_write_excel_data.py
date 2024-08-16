import openpyxl
"""
pip install openpyxl
"""


def read_content_from_excel(filename, sheet_name, cell_name):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet_name]
    cell = sheet[cell_name]
    print(cell.value)


#read_content_from_excel("test_excel_data.xlsx", sheet_name="Sheet1", cell_name="A2")
#read_content_from_excel("test_excel_data.xlsx", sheet_name="Sheet1", cell_name="A3")
#read_content_from_excel("test_excel_data.xlsx", sheet_name="Automation", cell_name="A2")


#for i in range(1, 5):
#   read_content_from_excel("test_excel_data.xlsx", sheet_name="Automation", cell_name=f"A{i}")

"""
Python
Selenium
Pytest
API Testing
"""


def read_all_rows_content_from_excel(filename, sheet_name):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet_name]
    max_row = sheet.max_row
    print(max_row)
    max_col = sheet.max_column
    print(max_col)
    cell_val = sheet.cell(1, 1).value
    print(cell_val)

    for row in range(1, max_row+1):
        for col in range(1, max_col+1):
            print(sheet.cell(row, col).value, end=" ")
        print()




read_all_rows_content_from_excel("test_excel_data.xlsx", "Sheet2")

def write_content_to_excel_file(filename, sheet_name, cell_name, data):
    wb = openpyxl.load_workbook(filename)
    sh_name = wb[sheet_name]
    cell_name = sh_name[cell_name]
    cell_name.value = data
    wb.save(filename)


#write_content_to_excel_file("test_excel_data.xlsx", "Sheet1", "B1", "India")

# st_names = ['Shreyash', 'Subham', 'Manjunath', 'Swetha', 'Vinoth']
#
# for i in range(len(st_names)):
#     write_content_to_excel_file("test_excel_data.xlsx", "Sheet1", f"C{i+1}", st_names[i])




